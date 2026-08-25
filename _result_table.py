
from ._mapi import MidasAPI
from ._model import Model
from typing import Literal
from ._mapi import _getUNIT
from ._mapi import _setUNIT
from ._view import View, ResultGraphic
from base64 import b64decode
# js_file = open('JSON_Excel Parsing\\test.json','r')

# print(js_file)
# js_json = json.load(js_file)

_SD_Type = Literal["X","Y","COMB"]
_forceType = Literal["KN", "N", "KGF", "TONF", "LBF", "KIPS"]
_lengthType = Literal["M", "CM", "MM", "FT", "IN"]
_numFormat = Literal["Fixed","Scientific","General"]
_resTable = Literal["REACTIONG","REACTIONL","DISPLACEMENTG","DISPLACEMENTL","TRUSSFORCE","TRUSSSTRESS"]
_Ref_Point_Type = Literal["Ground","Add Ground Motion","Another Node"]
_reactionType = Literal["Global", "Local", "SurfaceSpring"]
_dispdiaType = Literal["Global", "Local"]
_dispType = Literal["Accumulative", "Current", "Real"]
_plateforce = Literal["Global", "Local"]
_eigenOutputType = Literal["EigenVector","Eigenvalue_Analysis","Modal_Participation_Percent","Modal_Participation_Mass","Modal_Participation_Factor","Modal_Direction_Factor"]
_bucklingOutputType = Literal["BucklingVector","Buckling_Analysis"]
_EIGEN_SUBTABLE_MAP = {"Eigenvalue_Analysis":"EIGENVALUEANALYSIS","Modal_Participation_Percent":"MODALPARTICIPATIONMASSESPRINTOUT(1)","Modal_Participation_Mass":"MODALPARTICIPATIONMASSESPRINTOUT(2)","Modal_Participation_Factor":"MODALPARTICIPATIONFACTORPRINTOUT","Modal_Direction_Factor":"MODALDIRECTIONFACTORPRINTOUT","Buckling_Analysis":"BUCKLINGANALYSIS"}

def _convertColm2DataType(res_df):
    import pandas as pd
    
    # Column categorization using string matching
    str_colms = set([col for col in res_df.columns if any(x in col for x in ["Load", "Part", "Remark"])])
    int_colms1 = set([col for col in res_df.columns if col in ["Index", "Elem"]])
    int_colms2 = set([col for col in res_df.columns if "/Node" in col]) - int_colms1
    
    float_keywords = ["Axial", "Shear", "Torsion", "Moment", "FX", "FY", "FZ", "MX", "MY", "MZ", 
                      "DX", "DY", "DZ", "RX", "RY", "RZ", "Fx", "Fy", "Fz", "Mx", "My", "Mz", 
                      "Dx", "Dy", "Dz", "Rx", "Ry", "Rz", "Level", "Height", "Displacement", 
                      "Maximum", "Average", "Elements", "Drift", "Factor", "Frequency", "TRAN", 
                      "ROTN", "Period", "Tolerance", "Sig-", "Time", "Step"]
    float_colms = set([col for col in res_df.columns 
                       if any(x in col for x in float_keywords)]) - str_colms - int_colms1 - int_colms2
    
    # Type conversions
    res_type_df = res_df.copy()
    
    if str_colms:
        res_type_df[list(str_colms)] = res_type_df[list(str_colms)].astype(str)
    
    if int_colms1:
        res_type_df[list(int_colms1)] = res_type_df[list(int_colms1)].astype('int64')
    
    if int_colms2:
        res_type_df[list(int_colms2)] = res_type_df[list(int_colms2)].astype('int64')
    
    if float_colms:
        res_type_df[list(float_colms)] = res_type_df[list(float_colms)].astype('float32')
    
    # Handle "Node" column - convert if numeric, keep if alphanumeric
    if "Node" in res_type_df.columns:
        res_type_df["Node"] = res_type_df["Node"].apply(
            lambda x: int(x) if str(x).isdigit() else str(x)
        )
    
    return res_type_df


#---- INPUT: JSON -> OUTPUT : Data FRAME --------- ---------
def _JSToDF_ResTable(js_json, excelLoc, sheetName, cellLoc="A1", outputFormat='Pandas'):
    import pandas as pd
    
    if "SS_Table" not in js_json:
        if 'message' in js_json:
            print(f'⚠️  Error from API: {js_json["message"]}')
        else:
            print('⚠️  Error: "SS_Table" not found in the response JSON.')
        return pd.DataFrame()
        
    res_json = {}
    c = 0
    
    if "HEAD" not in js_json["SS_Table"] or "DATA" not in js_json["SS_Table"]:
        print('⚠️  Error: "HEAD" or "DATA" not found in "SS_Table".')
        return pd.DataFrame()
        
    for heading in js_json["SS_Table"]["HEAD"]:
        res_json[heading] = []
        for dat in js_json["SS_Table"]["DATA"]:
            try:
                res_json[heading].append(dat[c])
            except IndexError:
                pass
        c += 1

    res_df = pd.DataFrame(res_json)
    res_type_df = _convertColm2DataType(res_df)

    if excelLoc:
        _write_df_to_existing_excel(res_type_df, (excelLoc, sheetName, cellLoc))

    if outputFormat == 'JSON':
        return res_type_df.to_dict(orient='list')
    else:
        return res_type_df


#---- INPUT: JSON -> OUTPUT : Data FRAME --------- ---------
def _JSToDF_ResTable_TEXT(table_type, js_json, excelLoc, sheetName, cellLoc="A1", outputFormat='Pandas'):
    import pandas as pd
    
    if table_type not in js_json:
        if 'message' in js_json:
            print(f'⚠️  Error from API: {js_json["message"]}')
        else:
            print(f'⚠️  Error: "{table_type}" not found in the response JSON.')
        return pd.DataFrame()

    res_json = {}
    c = 0

    if "HEAD" not in js_json[table_type] or "DATA" not in js_json[table_type]:
        print(f'⚠️  Error: "HEAD" or "DATA" not found in "{table_type}".')
        return pd.DataFrame()

    for heading in js_json[table_type]["HEAD"]:
        res_json[heading] = []
        for dat in js_json[table_type]["DATA"]:
            try:
                res_json[heading].append(dat[c])
            except IndexError:
                pass
        c += 1

    res_df = pd.DataFrame(res_json)
    res_type_df = _convertColm2DataType(res_df)

    if excelLoc:
        _write_df_to_existing_excel(res_type_df, (excelLoc, sheetName, cellLoc))

    return res_type_df

def _format_mode_name(m):
    ''' Normalizes 1 / '1' / 'Mode 1' / 'Mode1' -> 'Mode1' '''
    s = str(m).strip().lower().replace("mode", "").strip()
    return f"Mode{s}"


def _format_modes(modes):
    if isinstance(modes, (list, tuple, set)):
        return [_format_mode_name(m) for m in modes]
    return [_format_mode_name(modes)]


#---- INPUT: JSON (Eigen result with SUB_TABLES) -> OUTPUT : Data FRAME ----
def _JSToDF_ResTable_Eigen(js_json, output, excelLoc, sheetName, cellLoc="A1", outputFormat='Pandas'):
    import pandas as pd

    if "SS_Table" not in js_json:
        if 'message' in js_json:
            print(f'⚠️  Error from API: {js_json["message"]}')
        else:
            print('⚠️  Error: "SS_Table" not found in the response JSON.')
        return pd.DataFrame()

    table_json = js_json["SS_Table"]

    if output == "EigenVector" or output == "BucklingVector":
        if "HEAD" not in table_json or "DATA" not in table_json:
            print('⚠️  Error: "HEAD" or "DATA" not found in "SS_Table".')
            return pd.DataFrame()

        res_json = {}
        c = 0
        for heading in table_json["HEAD"]:
            res_json[heading] = []
            for dat in table_json["DATA"]:
                try:
                    res_json[heading].append(dat[c])
                except IndexError:
                    pass
            c += 1

        res_df = pd.DataFrame(res_json)
        res_type_df = _convertColm2DataType(res_df)

    else:
        sub_key = _EIGEN_SUBTABLE_MAP.get(output)
        if sub_key is None:
            print(f'⚠️  Error: Unknown output type "{output}".')
            return pd.DataFrame()

        if "SUB_TABLES" not in table_json:
            print('⚠️  Error: "SUB_TABLES" not found in "SS_Table".')
            return pd.DataFrame()

        sub_table_data = None
        for sub_tab in table_json["SUB_TABLES"]:
            key_name = next(iter(sub_tab))
            key_norm = key_name.replace(" ", "").upper()
            if key_norm == sub_key or key_norm.startswith(sub_key):
                sub_table_data = sub_tab[key_name]
                break

        if sub_table_data is None:
            print(f'⚠️  Error: Sub-table for "{output}" not found in response.')
            return pd.DataFrame()

        if "HEAD" not in sub_table_data or "DATA" not in sub_table_data:
            print(f'⚠️  Error: "HEAD" or "DATA" not found in sub-table for "{output}".')
            return pd.DataFrame()

        res_json = _Head_Data_2_DF_JSON(sub_table_data["HEAD"], sub_table_data["DATA"])
        res_df = pd.DataFrame(res_json)
        res_type_df = _convertColm2DataType(res_df)

    if excelLoc:
        _write_df_to_existing_excel(res_type_df, (excelLoc, sheetName, cellLoc))

    return res_type_df





# #---- INPUT: JSON -> OUTPUT : Data FRAME --------- ---------
# def JSON2DF(json_data):
#     ''' CONVERTS JSON DATA WITH HEAD AND DATA KEYS TO POLARS DATAFRAME'''
#     # Check for SS_Table existence

#     import polars as pl

        
#     res_json = {}
#     c=0
    
#     # Check for HEAD and DATA existence
#     if "HEAD" not in json_data or "DATA" not in json_data:
#         print('⚠️  Error: "HEAD" or "DATA" not found in "SS_Table".')
#         return pl.DataFrame() # Return empty DataFrame
        
#     for heading in json_data["HEAD"]:
#         for dat in json_data["DATA"]:
#             try:
#                 res_json[heading].append(dat[c])
#             except:
#                 res_json[heading]=[]
#                 res_json[heading].append(dat[c])

#         c+=1

#     res_df = pl.DataFrame(res_json) # Final DF

#     res_type_df = _convertColm2DataType(res_df)

#     return(res_type_df)



def _Head_Data_2_DF_JSON(head,data):
    res_json = {}
    c=0
    headers = []
    for heading in head:
        if heading not in headers:
            headers.append(heading)
        elif f"{heading}_2" not in headers:
            headers.append(f"{heading}_2")
        elif f"{heading}_3" not in headers:
            headers.append(f"{heading}_3")
        elif f"{heading}_4" not in headers:
            headers.append(f"{heading}_4")  # Upto 4 repeated column names | Manually handled here

    for heading in headers:
        for dat in data:
            try:
                res_json[heading].append(dat[c])
            except:
                res_json[heading]=[]
                res_json[heading].append(dat[c])

        c+=1
    return res_json
    

def _JSToDF_UserDefined(tableName, js_json, summary, excelLoc, sheetName, cellLoc="A1", outputFormat='Pandas'):
    import pandas as pd
    
    if 'message' in js_json:
        print(f'⚠️  {tableName} table name does not exist.')
        return 'Check table name'
    
    if tableName not in js_json:
        print(f'⚠️  Error: Table "{tableName}" not found in API response.')
        return 'Check table name'

    if summary == 0:
        head = js_json[tableName]["HEAD"]
        data = js_json[tableName]["DATA"]
    elif summary > 0:
        try:
            sub_tab1 = js_json[tableName]["SUB_TABLES"][summary - 1]
            key_name = next(iter(sub_tab1))
            head = sub_tab1[key_name]["HEAD"]
            data = sub_tab1[key_name]["DATA"]
        except:
            print(' ⚠️  No Summary table exist')
            return 'No Summary table exist'

    res_json = _Head_Data_2_DF_JSON(head, data)
    res_df = pd.DataFrame(res_json)
    res_type_df = _convertColm2DataType(res_df)

    if excelLoc:
        _write_df_to_existing_excel(res_type_df, (excelLoc, sheetName, cellLoc))

    return res_type_df


    
def _write_df_to_existing_excel(res_df, existing_excel_input: list):
    import openpyxl
    from openpyxl.utils import column_index_from_string
    from openpyxl.styles import Font, PatternFill, Border, Side

    try:
        excel_path, sheet_name, start_cell = existing_excel_input
        if not all([excel_path, sheet_name, start_cell]):
            print("⚠️  `existing_excel_input` has empty values. Skipping update.")
            return

        try:
            wb = openpyxl.load_workbook(excel_path)
        except:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            print(f"      ⚠️  Sheet '{sheet_name}' not found in {excel_path}. Creating new sheet.")
            ws = wb.create_sheet(sheet_name)
        
        if start_cell == "end":
            nRow = ws.max_row
            if nRow == 1:
                nRow = -1
            start_row_str = nRow + 2
            start_col = ws.min_column
        else:
            start_col_let = ''.join(filter(str.isalpha, start_cell))
            start_row_str = ''.join(filter(str.isdigit, start_cell))
            start_col = column_index_from_string(start_col_let)
            
        start_row = int(start_row_str)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="000000")
        thin = Side(style="thin")

        # Write header
        headers = res_df.columns
        for c_idx, header in enumerate(headers):
            cell = ws.cell(row=start_row, column=start_col + c_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill


        for r_idx, row_data in enumerate(res_df.itertuples(index=False, name=None)):
            for c_idx, cell_value in enumerate(row_data):
                cell = ws.cell(row=start_row + 1 + r_idx, column=start_col + c_idx, value=cell_value)
                cell.border = Border(bottom=thin)
        
        wb.save(excel_path)
        wb.close()
        print(f"      ✅ Updated excel file: {excel_path} | Sheet: {sheet_name} | Cell: {start_cell} |")

    except Exception as e:
        print(f"⚠️  Error writing to existing Excel file: {e}")

def _changeUNITandGetData(js_dat,force_unit,len_unit,jsonloc,keyName):
    # currUNIT = _getUNIT()
    Model.units(force=force_unit,length=len_unit)
    ss_json = MidasAPI("POST","/post/table",js_dat)
    # _setUNIT(currUNIT)
    if jsonloc:
        if "SS_Table" in ss_json:
            ss_json[keyName] = ss_json.pop("SS_Table")
            _saveJSON(ss_json,jsonloc)
            ss_json["SS_Table"] = ss_json.pop(keyName)
        else:
            _saveJSON(ss_json,jsonloc)
    return ss_json

def _changeUNITandGetDataText(js_dat,force_unit,len_unit,jsonloc,keyName):
    Model.units(force=force_unit,length=len_unit)
    ss_json = MidasAPI("POST","/post/TEXT",js_dat)
    if jsonloc:
        if "SS_Table" in ss_json:
            ss_json[keyName] = ss_json.pop("SS_Table")
            _saveJSON(ss_json,jsonloc)
            ss_json["SS_Table"] = ss_json.pop(keyName)
        else:
            _saveJSON(ss_json,jsonloc)
    return ss_json

def _keys2JSON(keys):
    if isinstance(keys,list):
        if keys!=[]:
            out_js = {"KEYS": keys}
    elif isinstance(keys,str):
        out_js = {"STRUCTURE_GROUP_NAME": keys}
    return out_js


def _saveJSON(jsonData,fileLocation = "jsData.json"):
        import json
        with open(fileLocation, "w", encoding="utf-8") as f:
            json.dump(jsonData, f, indent=4, ensure_ascii=False)


def _case2name(s):
    if isinstance(s,str):
        return f'{s.split("(")[0]} LCase'
    if isinstance(s,list):
        n = len(s)
        if n==0:
            return f"Table"
        return f"{n} LCases"

def _generate(table_type,keys,loadcase,components,cs_stage,options):
    js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": table_type,
                    "STYLES": options.Style,
                }
            }

    if keys: js_dat["Argument"]['NODE_ELEMS'] = _keys2JSON(keys)
    if loadcase: js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase
    if components != ['all']: js_dat["Argument"]['COMPONENTS'] = components

    if cs_stage !=[]:
        if cs_stage == 'all' or cs_stage == ['all']:
            js_dat["Argument"]['OPT_CS'] = True
        else:
            js_dat["Argument"]['OPT_CS'] = True
            js_dat["Argument"]['STAGE_STEP'] = cs_stage
   
    return js_dat

class TableOptions:
    FORCE_UNIT = 'KN'
    LEN_UNIT = 'M'
    NUM_FORMAT = 'Fixed'
    DECIMAL_PLACE = 5
    # JSON_OUTPUT_LOC = None
    EXCEL_FILE_LOC = None
    EXCEL_SHEET_NAME = None
    EXCEL_CELL_POS = "end"
    OUTPUT_FORMAT = 'Pandas'

    def __init__(self,force_unit:_forceType=None,len_unit:_lengthType=None,num_format:_numFormat=None,decimal_place:int=None,
                 JSONFileLoc=None,ExcelFileLoc=None , ExcelSheetName = None,ExcelCellPos = None , outputFormat = None):
        
        # existing_excel_input -> excel file , sheet , cell

        '''
        Table Options
        
        :param force_unit: Enter force unit - "KN", "N", "KGF", "TONF", "LBF", "KIPS"
        :param len_unit: Enter length unit - "M", "CM", "MM", "FT", "IN"
        :param num_format: Enter number format - "Fixed","Scientific","General"
        :param decimal_place: Number of decimal places for result output
        '''
        self.FORCE_UNIT = force_unit or TableOptions.FORCE_UNIT
        self.LEN_UNIT = len_unit or TableOptions.LEN_UNIT
        self.NUM_FORMAT = num_format or TableOptions.NUM_FORMAT
        self.DECIMAL_PLACE = decimal_place or TableOptions.DECIMAL_PLACE
        # self.JSON_OUTPUT_LOC = JSONLoc or TableOptions.JSON_OUTPUT_LOC
        self.JSON_FILE_LOC = JSONFileLoc
        self.EXCEL_FILE_LOC = ExcelFileLoc or TableOptions.EXCEL_FILE_LOC
        self.EXCEL_SHEET_NAME = ExcelSheetName or TableOptions.EXCEL_SHEET_NAME
        self.EXCEL_CELL_POS = ExcelCellPos or TableOptions.EXCEL_CELL_POS
        self.OUTPUT_FORMAT = outputFormat or TableOptions.OUTPUT_FORMAT

    @property
    def Style(self):
        '''rrr'''
        if self.NUM_FORMAT == 'Fixed':
            js = {"FORMAT" : "Fixed" , "PLACE":self.DECIMAL_PLACE}
        else:
            js = {"FORMAT" : self.NUM_FORMAT}
        return js
    
    @property
    def Unit(self):
        ''' rr '''
        return {"FORCE": self.FORCE_UNIT, "DIST": self.LEN_UNIT }
    

    def __str__(self):
        return str(self.__dict__)

class Result :

    # ---------- Result TABLE (For ALL TABLES)------------------------------    

    class TABLE :
        '''
        Extracts tabular result from MIDAS CIVIL NX
        '''

        def __new__(cls,tabletype:_resTable,keys=[],loadcase:list=[],cs_stage=[],options:TableOptions=None):
            '''
                TableType : REACTIONG | REACTIONL | DISPLACEMENTG | DISPLACEMENTL | TRUSSFORCE | TRUSSSTRESS
                Keys : List{int} -> Element/ Node IDs  |  str -> Structure Group Name
                Loadcase : Loadcase/Combination name followed by type. eg. DeadLoad(ST)
            '''
            instance = super().__new__(cls)
            return instance._dispatch(tabletype, keys,loadcase,cs_stage,options)
        
        @classmethod
        def _dispatch(cls,tabletype, keys,loadcase,cs_stage,options):
            if options == None : options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"{tabletype} {_case2name(loadcase)}"

            js_dat = _generate(tabletype,keys,loadcase,[],cs_stage,options)

            ResultJSON = _changeUNITandGetData(js_dat,options.FORCE_UNIT,options.LEN_UNIT,options.JSON_FILE_LOC,tabletype)
            polarDF = _JSToDF_ResTable(ResultJSON,options.EXCEL_FILE_LOC,sheetName,options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF
        
            # ---------- User defined TABLE (Dynamic Report Table) ------------------------------
        @staticmethod
        def UserDefinedTable(tableName:str, summary=0,options:TableOptions=None):
            if options == None : options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"{tableName} Table"
            js_dat = {
                "Argument": {
                    "TABLE_NAME": tableName,
                    "STYLES": options.Style
                }
            }



            ResultJSON = _changeUNITandGetData(js_dat,options.FORCE_UNIT,options.LEN_UNIT,options.JSON_FILE_LOC,tableName)
            polarDF = _JSToDF_UserDefined(tableName,ResultJSON,summary,options.EXCEL_FILE_LOC,sheetName,options.EXCEL_CELL_POS)
            return polarDF

        
        # ---------- LIST ALL USER DEFINED TABLE ------------------------------
        @staticmethod
        def UserDefinedTables_list():
            ''' Print all the User defined table names '''
            ss_json = MidasAPI("GET","/ope/UTBLTYPES",{})
            table_name =[]
            try:
                for tabName in ss_json['UTBLTYPES']:
                    table_name.append(tabName)
                
                print('Available user-defined tables in Civil NX are : ')
                print(*table_name,sep=' , ')
            except:
                print(' ⚠️  There are no user-defined tables in Civil NX')

            return table_name

        @staticmethod
        def Reaction(keys=[], loadcase:list=[], components=['all'],
                     cs_stage=[],
                     type:_reactionType="Global",options:TableOptions= None):
            '''
            Fetches Reaction result tables (Global, Local, or Surface Spring).
            
            Args:
                keys (list/str): List of Node IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["DL(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                type (str): Reaction type. "Global", "Local", or "SurfaceSpring"
                options : table option
            '''
            if options == None : options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"Reaction {_case2name(loadcase)}"

            table_type_map = {
                "Global": "REACTIONG",
                "Local": "REACTIONL",
                "SurfaceSpring": "REACTIONLSURFACESPRING"
            }
            table_type = table_type_map.get(type.capitalize(), "REACTIONG") # Default to Global
            js_dat = _generate(table_type,keys,loadcase,components,cs_stage,options)

            ResultJSON = _changeUNITandGetData(js_dat,options.FORCE_UNIT,options.LEN_UNIT,options.JSON_FILE_LOC,table_type)
            polarDF = _JSToDF_ResTable(ResultJSON,options.EXCEL_FILE_LOC,sheetName,options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF

        @staticmethod
        def Displacement(keys=[], loadcase:list=[], components=['all'], 
                         cs_stage=[],
                         type:_dispdiaType ="Global",
                         displacement_type:_dispType="Accumulative",
                         options:TableOptions=None):
            '''
            Fetches Displacement result tables (Global or Local).
            
            Args:
                keys (list/str): List of Node IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["Self(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                type (str): Displacement type. "Global" or "Local".
                displacement_type (str): "Accumulative", "Current", or "Real".
                options : Table options
            '''
            if options == None : options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"Displacement {_case2name(loadcase)}"

            table_type_map = {
                "Global": "DISPLACEMENTG",
                "Local": "DISPLACEMENTL"
            }
            table_type = table_type_map.get(type.capitalize(), "DISPLACEMENTG") 

            js_dat = _generate(table_type,keys,loadcase,components,cs_stage,options)

            if displacement_type in ["Accumulative", "Current", "Real"]:
                js_dat["Argument"]["DISP_OPT"] = displacement_type

            ResultJSON = _changeUNITandGetData(js_dat,options.FORCE_UNIT,options.LEN_UNIT,options.JSON_FILE_LOC,table_type)
            polarDF = _JSToDF_ResTable(ResultJSON,options.EXCEL_FILE_LOC,sheetName,options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF
    
        @staticmethod
        def TrussForce(keys=[], loadcase:list=[], components=['all'], 
                       cs_stage=[], options:TableOptions=None):
            '''
            Fetches Truss Force result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["DL(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"TrussForce {_case2name(loadcase)}"
            
            table_type = "TRUSSFORCE"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF

        @staticmethod
        def TrussStress(keys=[], loadcase:list=[], components=['all'], 
                        cs_stage=[], options:TableOptions=None):
            '''
            Fetches Truss Stress result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["DL(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"TrussStress {_case2name(loadcase)}"
            
            table_type = "TRUSSSTRESS"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF

        @staticmethod
        def BeamForce(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                      components=['all'], cs_stage=[], options:TableOptions=None):
            '''
            Fetches standard Beam Force result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["Selfweight(ST)"].
                parts (list): Element parts: ["PartI", "Part1/4", "PartJ", etc.].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"BeamForce {_case2name(loadcase)}"
            
            table_type = "BEAMFORCE"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            if parts:
                js_dat["Argument"]["PARTS"] = parts

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF

        @staticmethod
        def BeamForce_VBM(keys=[], loadcase:list=[], items=['all'], parts=["PartI", "PartJ"], 
                          components=['all'], cs_stage=[], options:TableOptions=None):
            '''
            Fetches Beam Force (View by Max Value) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["STLENV_STR(CB:max)"].
                items (list): Items to display: ["Axial", "Shear-y", "Moment-z", etc.].
                parts (list): Element parts: ["PartI", "Part1/4", "PartJ", etc.].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"BeamForceVBM {_case2name(loadcase)}"
            
            table_type = "BEAMFORCEVBM"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            if parts:
                js_dat["Argument"]["PARTS"] = parts
            
            if items != ['all']:
                js_dat["Argument"]['ITEM_TO_DISPLAY'] = items

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF

        @staticmethod
        def BeamForce_StaticPrestress(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                                      components=['all'], options:TableOptions=None):
            '''
            Fetches Beam Force (Static Prestress) result tables.
            Note: Construction Stage options are not applicable to this table type.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["Prestress(ST)"].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                components (list): Table components to include. Defaults to ['all'].
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"BeamForceSTP {_case2name(loadcase)}"
            
            table_type = "BEAMFORCESTP"
            
            # Note: cs_stage is intentionally omitted for this type
            js_dat = _generate(table_type, keys, loadcase, components, [], options)
            
            if parts:
                js_dat["Argument"]["PARTS"] = parts

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF

        @staticmethod
        def BeamStress(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                       components=['all'], cs_stage=[], options:TableOptions=None):
            '''
            Fetches standard Beam Stress result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["Selfweight(ST)"].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"BeamStress {_case2name(loadcase)}"
            
            table_type = "BEAMSTRESS"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            if parts:
                js_dat["Argument"]["PARTS"] = parts

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF

        @staticmethod
        def BeamStress_VBM(keys=[], loadcase:list=[], items=['all'], parts=["PartI", "PartJ"], 
                           components=['all'], cs_stage=[], options:TableOptions=None):
            '''
            Fetches Beam Stress (View by Max Value) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["STLENV_SER(CB:max)"].
                items (list): Items to display: ["Axial", "Shear-y", "Bend(+y)", etc.].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"BeamStressVBM {_case2name(loadcase)}"
            
            table_type = "BEAMSTRESSVBM"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            if parts:
                js_dat["Argument"]["PARTS"] = parts
            
            if items != ['all']:
                js_dat["Argument"]['ITEM_TO_DISPLAY'] = items

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF

        @staticmethod
        def BeamStress_7DOF(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                            section_position=['Max'], components=['all'], 
                            cs_stage=[], options:TableOptions=None):
            '''
            Fetches Beam Stress (7th DOF) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["EccentricLoads(ST)"].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                section_position (list): Section positions: ["Pos-1", "Pos-4", "Max", etc.].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"BeamStress7DOF {_case2name(loadcase)}"
            
            table_type = "BEAMSTRESS7DOF"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            if parts:
                js_dat["Argument"]["PARTS"] = parts
            
            if section_position:
                js_dat["Argument"]["SECTION_POSITION"] = section_position

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF

        @staticmethod
        def BeamStress_PSC(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                           section_position=['All'], components=['all'], 
                           cs_stage=[], options:TableOptions=None):
            '''
            Fetches Beam Stress (PSC) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["Selfweight(ST)"].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                section_position (list): Section positions: ["Pos-1", "Pos-10", "Max", "Min", "All"].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"BeamStressPSC {_case2name(loadcase)}"
            
            table_type = "BEAMSTRESSPSC"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            if parts:
                js_dat["Argument"]["PARTS"] = parts
            
            if section_position:
                js_dat["Argument"]["SECTION_POSITION"] = section_position

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF

        @staticmethod
        def BeamStress_7DOF_PSC(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                                section_position=['All'], components=['all'], 
                                cs_stage=[], options:TableOptions=None):
            '''
            Fetches Beam Stress (7th DOF PSC) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["EccentricLoads(ST)"].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                section_position (list): Section positions: ["Pos-1", "Pos-10", "Max", "Min", "All"].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"BeamStress7DOFPSC {_case2name(loadcase)}"
            
            table_type = "BEAMSTRESS7DOFPSC"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            if parts:
                js_dat["Argument"]["PARTS"] = parts
            
            if section_position:
                js_dat["Argument"]["SECTION_POSITION"] = section_position

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF

        @staticmethod
        def PlateForce(keys=[], loadcase:list=[], components=['all'], 
                       cs_stage=[], avg_nodal_result=False,
                       type:str="Local", options:TableOptions=None):
            '''
            Fetches Plate Force (Local or Global) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["DL(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                avg_nodal_result (bool): Option to average nodal results.
                type (str): Plate Force type. "Local" or "Global".
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"PlateForce{type} {_case2name(loadcase)}"
            
            table_type_map = {
                "Local": "PLATEFORCEL",
                "Global": "PLATEFORCEG"
            }
            table_type = table_type_map.get(type.capitalize(), "PLATEFORCEL")
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            if avg_nodal_result:
                js_dat["Argument"]["AVERAGE_NODAL_RESULT"] = True

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF

        @staticmethod
        def BeamStress_Equivalent(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                                  section_position=['Maximum'], components=['all'], 
                                  cs_stage=[], options:TableOptions=None):
            '''
            Fetches Beam Stress (Equivalent) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["Selfweight(ST)"].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                section_position (list): Section positions: ["Maximum", "1", "12", etc.].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"BeamStressEq {_case2name(loadcase)}"
            
            table_type = "BEAMSTRESSDETAIL"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            if parts:
                js_dat["Argument"]["PARTS"] = parts
            
            if section_position:
                js_dat["Argument"]["SECTION_POSITION"] = section_position

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF

        @staticmethod
        def PlateForce_UnitLength(keys=[], loadcase:list=[], components=['all'], 
                                cs_stage=[], avg_nodal_result=False,
                                node_flag_center=False, node_flag_nodes=True,
                                type:_plateforce="Local", options:TableOptions=None):
            '''
            Fetches Plate Force (Unit Length) for Local or UCS coordinates.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["DL(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                avg_nodal_result (bool): Option to average nodal results.
                node_flag_center (bool): Retrieve results at the center of the plate.
                node_flag_nodes (bool): Retrieve results at the nodes of the plate.
                type (str): Plate Force type. "Local" or "Global"
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"PlateForceUL{type} {_case2name(loadcase)}"
            
            table_type_map = {
                "Local": "PLATEFORCEUL",
                "Global": "PLATEFORCEUG" 
            }
            table_type = table_type_map.get(type.capitalize(), "PLATEFORCEUL")
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            js_dat["Argument"]["NODE_FLAG"] = {
                "CENTER": node_flag_center,
                "NODES": node_flag_nodes
            }
            
            if avg_nodal_result:
                js_dat["Argument"]["AVERAGE_NODAL_RESULT"] = True

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF

        @staticmethod
        def PlateForce_UnitLength_VBM(keys=[], loadcase:list=[], items=['all'], 
                                      components=['all'], cs_stage=[], 
                                      avg_nodal_result=False,
                                      node_flag_center=False, node_flag_nodes=True,
                                      type:_plateforce="Local", options:TableOptions=None):
            '''
            Fetches Plate Force (Unit Length, View by Max Value) for Local or UCS coordinates.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["STLENV_STR(CB:max)"].
                items (list): Items to display: ["Fxx", "Fyy", "Mxx", etc.].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                avg_nodal_result (bool): Option to average nodal results.
                node_flag_center (bool): Retrieve results at the center of the plate.
                node_flag_nodes (bool): Retrieve results at the nodes of the plate.
                type (str): Plate Force type. "Local" or "Global"
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"PlateForceULVBM{type} {_case2name(loadcase)}"
            
            table_type_map = {
                "Local": "PLATEFORCEULVBM",
                "Global": "PLATEFORCEUGVBM" 
            }
            table_type = table_type_map.get(type.capitalize(), "PLATEFORCEULVBM")
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            js_dat["Argument"]["NODE_FLAG"] = {
                "CENTER": node_flag_center,
                "NODES": node_flag_nodes
            }
            
            if items != ['all']:
                js_dat["Argument"]['ITEM_TO_DISPLAY'] = items
            
            if avg_nodal_result:
                js_dat["Argument"]["AVERAGE_NODAL_RESULT"] = True

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF

        @staticmethod
        def PlateForce_UnitLength_WA(keys=[], loadcase:list=[], components=['all'], 
                                   cs_stage=[], avg_nodal_result=False,
                                   node_flag_center=False, node_flag_nodes=True,
                                   options:TableOptions=None):
            '''
            Fetches Plate Force (Unit Length, W-A Moment) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["DL(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                avg_nodal_result (bool): Option to average nodal results.
                node_flag_center (bool): Retrieve results at the center of the plate.
                node_flag_nodes (bool): Retrieve results at the nodes of the plate.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"PlateForceWA {_case2name(loadcase)}"
            
            table_type = "PLATEFORCEWA"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            js_dat["Argument"]["NODE_FLAG"] = {
                "CENTER": node_flag_center,
                "NODES": node_flag_nodes
            }
            
            if avg_nodal_result:
                js_dat["Argument"]["AVERAGE_NODAL_RESULT"] = True

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF
        
        @staticmethod
        def Tendon_Loss(tdn_group="", cs_stage="",
                                   options:TableOptions=None):
            '''
            Fetches Tendon Loss result tables.
            
            Args:
                tdn_group (str): Tendon Name or a Tendon Group Name.
                cs_stage (str): Construction Stage Name.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"Tendon Loss - {tdn_group}"
            
            table_type = "TNDN_LOSS_FORCE"
            
            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": table_type,
                    "STYLES": options.Style,
                    "ADDITIONAL": {
                        "SET_TENDON_PARAMS": {
                            "TENDON_GROUP": tdn_group,
                            "STAGE": cs_stage
                        }
                    }
                }
            }

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF
        
                
        @staticmethod
        def Story_Displacement(SD_type:_SD_Type = "X",
                                loadcase:list=[], 
                                options:TableOptions=None):
            '''
            Fetches story displacement - story result tables.
            
            Args:
                SD_type (str): "X","Y" or "COMB
                loadcase (list): List of load case names, e.g., ["Selfweight(ST)"].
                options : Table options
            '''
            if options == None: options = TableOptions()

            if SD_type == "X":
                table_type = "STORY_DISPLACEMENT_X"
            elif SD_type == "Y":
                table_type = "STORY_DISPLACEMENT_Y"
            else:
                table_type = "STORY_DISPLACEMENT_COMB"

            sheetName = options.EXCEL_SHEET_NAME or f"{table_type} {_case2name(loadcase)}"
            
            js_dat = _generate(table_type,None,loadcase,None,None, options)
            

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF
        
        @staticmethod
        def Story_Drift(SD_type:_SD_Type = "X",
                        loadcase:list=[], 
                        allowable_ratio:float = 0.0015,
                        options:TableOptions=None):
            '''
            Fetches story drift -    story result tables.
            
            Args:
                SD_type (str): "X","Y" or "COMB"
                loadcase (list): List of load case names, e.g., ["Selfweight(ST)"].
                allowable_ratio (float): Allowable Story Drift Ratio
                options : Table options
            '''
            if options == None: options = TableOptions()

            if SD_type == "X":
                table_type = "STORY_DRIFT_X"
            elif SD_type == "Y":
                table_type = "STORY_DRIFT_Y"
            else:
                table_type = "STORY_DRIFT_COMB"

            sheetName = options.EXCEL_SHEET_NAME or f"{table_type} {_case2name(loadcase)}"
            
            js_dat = _generate(table_type,None,loadcase,None,None, options)
            js_dat["Argument"]['SET_STORY_DRIFT_PARAMS'] = {}
            js_dat["Argument"]['SET_STORY_DRIFT_PARAMS']['RESPONSE_MOD_FACTOR_CHECK']  = True
            js_dat["Argument"]['SET_STORY_DRIFT_PARAMS']['RESPONSE_MOD_FACTOR_VALUE']  = 1 
            js_dat["Argument"]['SET_STORY_DRIFT_PARAMS']['SCALE_FACTOR_VALUE']  = 1
            js_dat["Argument"]['SET_STORY_DRIFT_PARAMS']['ALLOWABLE_RATIO']  = allowable_ratio

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF

        @staticmethod
        def HoH_Stress(keys=[],Stress_option:str = "Local", node_flag_center=False, node_flag_nodes=True, 
                       components=['all'], cs_stage=[], options:TableOptions=None):
            '''
            Fetches Heat of Hydration Stress result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                Stress_option (str): Stress Option for "Local" or "Global".
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"Heat of Hydration Stress ({Stress_option})"

            if Stress_option == "Local":
                table_type = "HEAT_HYDR_STRESS_L"
            else:
                table_type = "HEAT_HYDR_STRESS_G"

            js_dat = _generate(table_type, keys, ['all'], components, cs_stage, options)

            js_dat["Argument"].pop("LOAD_CASE_NAMES")

            js_dat["Argument"]["NODE_FLAG"] = {
                "CENTER": node_flag_center,
                "NODES": node_flag_nodes
            }

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF

        @staticmethod
        def HoH_Temperature(keys=[],components=['all'], cs_stage=[], options:TableOptions=None):
            '''
            Fetches Heat of Hydration Temperature result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"Heat of Hydration Temperature"

            table_type = "HEAT_HYDR_TEMPERATURE"

            js_dat = _generate(table_type, keys, ['all'], components, cs_stage, options)

            js_dat["Argument"].pop("LOAD_CASE_NAMES")

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF

        @staticmethod
        def HoH_Displacement(keys=[],components=['all'], cs_stage=[], options:TableOptions=None):
            '''
            Fetches Heat of Hydration Displacement result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"Heat of Hydration Displacement"

            table_type = "HEAT_HYDR_DISPLACEMENT"

            js_dat = _generate(table_type, keys, ['all'], components, cs_stage, options)

            js_dat["Argument"].pop("LOAD_CASE_NAMES")

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF

        @staticmethod
        def HoH_Tensile_Stress(keys=[],components=['all'], cs_stage=[], options:TableOptions=None):
            '''
            Fetches Heat of Hydration Tensile Stress result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"Heat of Hydration Tensile Stress"

            table_type = "HEAT_HYDR_TENS_STRESS"

            js_dat = _generate(table_type, keys, ['all'], components, cs_stage, options)

            js_dat["Argument"].pop("LOAD_CASE_NAMES")

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF

        @staticmethod
        def HoH_Pipe_Node_Temperature(keys=[],components=['all'], cs_stage=[], options:TableOptions=None):
            '''
            Fetches Heat of Hydration Pipe Cooling Nodal Temperature Result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"Heat of Hydration Pipe Cooling Nodal Temperature"

            table_type = "HEAT_HYDR_PIPE_NODE_TEMP"

            js_dat = _generate(table_type, keys, ['all'], components, cs_stage, options)

            js_dat["Argument"].pop("LOAD_CASE_NAMES")

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS, options.OUTPUT_FORMAT)
            return polarDF
        
        @staticmethod
        def TH_Disp(th_case: list,keys=[],
                    step_from: float = 0, step_to: float = 1, step_interval: int = 1,
                    components=['all'],
                    ref_pt: _Ref_Point_Type = "Ground", anr_node: int = None,
                    options: TableOptions = None):
            '''
            Fetches Time History Displacement (Node) result tables.
            Args:
                keys (list/str): List of Node IDs or a Structure Group Name.
                th_case (list): Time history load case names, e.g. ["Elcent"].
                step_from (float): Start time.
                step_to (float): End time.
                step_interval (int): Time interval (STEPS).
                components (list): Table components to include. Defaults to ['all'].
                ref_pt (str): Reference point method 1 - "Ground" or "AddGroundMotion".
                            Ignored if `anr_node` is provided.
                anr_node (int): Reference point method 2 - another node number.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"TH_Disp {_case2name(th_case)}"
            table_type = "TH_DISP"
            js_dat = _generate(table_type, keys, th_case, components, [], options)

            js_dat["Argument"]["TEXT_TYPE"] = js_dat["Argument"].pop("TABLE_TYPE")

            if "LOAD_CASE_NAMES" in js_dat["Argument"]:
                js_dat["Argument"]["TH_CASE_NAME"] = js_dat["Argument"].pop("LOAD_CASE_NAMES")

            js_dat["Argument"]["STEP"] = {
                "FROM": step_from,
                "TO": step_to,
                "STEPS": step_interval
            }

            if anr_node is not None:
                js_dat["Argument"]["ANR_NODE"] = anr_node
            elif ref_pt:
                js_dat["Argument"]["REF_PT"] = ref_pt
                
            js_dat["Argument"].pop("TABLE_NAME", None)
            ResultJSON = _changeUNITandGetDataText(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable_TEXT(table_type,ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF
        
        @staticmethod
        def TH_Velocity(th_case: list,keys=[],
                        step_from: float = 0, step_to: float = 1, step_interval: int = 1,
                        components=['all'],
                        ref_pt: _Ref_Point_Type = "Ground", anr_node: int = None,
                        options: TableOptions = None):
            '''
            Fetches Time History Velocity (Node) result tables.
            Args:
                keys (list/str): List of Node IDs or a Structure Group Name.
                th_case (list): Time history load case names, e.g. ["Elcent"].
                step_from (float): Start time.
                step_to (float): End time.
                step_interval (int): Time interval (STEPS).
                components (list): Table components to include. Defaults to ['all'].
                ref_pt (str): Reference point method 1 - "Ground" or "AddGroundMotion".
                            Ignored if `anr_node` is provided.
                anr_node (int): Reference point method 2 - another node number.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"TH_Velocity {_case2name(th_case)}"
            table_type = "TH_VELOCITY"
            js_dat = _generate(table_type, keys, th_case, components, [], options)

            js_dat["Argument"]["TEXT_TYPE"] = js_dat["Argument"].pop("TABLE_TYPE")

            if "LOAD_CASE_NAMES" in js_dat["Argument"]:
                js_dat["Argument"]["TH_CASE_NAME"] = js_dat["Argument"].pop("LOAD_CASE_NAMES")

            js_dat["Argument"]["STEP"] = {
                "FROM": step_from,
                "TO": step_to,
                "STEPS": step_interval
            }

            if anr_node is not None:
                js_dat["Argument"]["ANR_NODE"] = anr_node
            elif ref_pt:
                js_dat["Argument"]["REF_PT"] = ref_pt

            js_dat["Argument"].pop("TABLE_NAME", None)
            ResultJSON = _changeUNITandGetDataText(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable_TEXT(table_type, ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def TH_Acceleration(th_case: list,keys=[],
                            step_from: float = 0, step_to: float = 1, step_interval: int = 1,
                            components=['all'],
                            ref_pt: _Ref_Point_Type = "Ground", anr_node: int = None,
                            options: TableOptions = None):
            '''
            Fetches Time History Acceleration (Node) result tables.
            Args:
                keys (list/str): List of Node IDs or a Structure Group Name.
                th_case (list): Time history load case names, e.g. ["Elcent"].
                step_from (float): Start time.
                step_to (float): End time.
                step_interval (int): Time interval (STEPS).
                components (list): Table components to include. Defaults to ['all'].
                ref_pt (str): Reference point method 1 - "Ground" or "AddGroundMotion".
                            Ignored if `anr_node` is provided.
                anr_node (int): Reference point method 2 - another node number.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"TH_Accel {_case2name(th_case)}"
            table_type = "TH_ACCEL"
            js_dat = _generate(table_type, keys, th_case, components, [], options)

            js_dat["Argument"]["TEXT_TYPE"] = js_dat["Argument"].pop("TABLE_TYPE")

            if "LOAD_CASE_NAMES" in js_dat["Argument"]:
                js_dat["Argument"]["TH_CASE_NAME"] = js_dat["Argument"].pop("LOAD_CASE_NAMES")

            js_dat["Argument"]["STEP"] = {
                "FROM": step_from,
                "TO": step_to,
                "STEPS": step_interval
            }

            if anr_node is not None:
                js_dat["Argument"]["ANR_NODE"] = anr_node
            elif ref_pt:
                js_dat["Argument"]["REF_PT"] = ref_pt

            js_dat["Argument"].pop("TABLE_NAME", None)
            ResultJSON = _changeUNITandGetDataText(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable_TEXT(table_type, ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF
        
        @staticmethod
        def TH_BeamForce(th_case: list,keys=[],parts=["PartI", "PartJ"],
                    step_from: float = 0, step_to: float = 1, step_interval: int = 1,
                    components=['all'],
                    options: TableOptions = None):
            '''
            Fetches Time History BeamForce result tables.
            Args:
                keys (list/str): List of Beam Element IDs.
                th_case (list): Time history load case names, e.g. ["Elcent"].
                step_from (float): Start time.
                step_to (float): End time.
                step_interval (int): Time interval (STEPS).
                components (list): Table components to include. Defaults to ['all'].
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"TH_BeamForce {_case2name(th_case)}"
            table_type = "TH_BEAMFORCE"
            js_dat = _generate(table_type, keys, th_case, components, [], options)

            js_dat["Argument"]["TEXT_TYPE"] = js_dat["Argument"].pop("TABLE_TYPE")
            
            if parts:
                js_dat["Argument"]["PARTS"] = parts
            
            if "LOAD_CASE_NAMES" in js_dat["Argument"]:
                js_dat["Argument"]["TH_CASE_NAME"] = js_dat["Argument"].pop("LOAD_CASE_NAMES")

            js_dat["Argument"]["STEP"] = {
                "FROM": step_from,
                "TO": step_to,
                "STEPS": step_interval
            }
                
            js_dat["Argument"].pop("TABLE_NAME", None)
            ResultJSON = _changeUNITandGetDataText(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable_TEXT(table_type,ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def TH_TrussForce(th_case: list,keys=[],
                    step_from: float = 0, step_to: float = 1, step_interval: int = 1,
                    components=['all'],
                    options: TableOptions = None):
            '''
            Fetches Time History TrussForce result tables.
            Args:
                keys (list/str): List of Truss Element IDs.
                th_case (list): Time history load case names, e.g. ["Elcent"].
                step_from (float): Start time.
                step_to (float): End time.
                step_interval (int): Time interval (STEPS).
                components (list): Table components to include. Defaults to ['all'].
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"TH_TrussForce {_case2name(th_case)}"
            table_type = "TH_TRUSSFORCE"
            js_dat = _generate(table_type, keys, th_case, components, [], options)

            js_dat["Argument"]["TEXT_TYPE"] = js_dat["Argument"].pop("TABLE_TYPE")

            if "LOAD_CASE_NAMES" in js_dat["Argument"]:
                js_dat["Argument"]["TH_CASE_NAME"] = js_dat["Argument"].pop("LOAD_CASE_NAMES")

            js_dat["Argument"]["STEP"] = {
                "FROM": step_from,
                "TO": step_to,
                "STEPS": step_interval
            }
                
            js_dat["Argument"].pop("TABLE_NAME", None)
            ResultJSON = _changeUNITandGetDataText(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable_TEXT(table_type,ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF


        @staticmethod
        def TH_PlaneStressForce(th_case: list,keys=[],parts: list=["PartI", "PartJ","PartK","PartL"],
                    step_from: float = 0, step_to: float = 1, step_interval: int = 1,
                    components=['all'],
                    options: TableOptions = None):
            '''
            Fetches Time History PlaneStressForce result tables.
            Args:
                keys (list/str): List of Plane Stress Element IDs.
                th_case (list): Time history load case names, e.g. ["Elcent"].
                step_from (float): Start time.
                step_to (float): End time.
                step_interval (int): Time interval (STEPS).
                components (list): Table components to include. Defaults to ['all'].
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"TH_PlaneStressForce {_case2name(th_case)}"
            table_type = "TH_PLANE_STRESS_FORCE"
            js_dat = _generate(table_type, keys, th_case, components, [], options)

            js_dat["Argument"]["TEXT_TYPE"] = js_dat["Argument"].pop("TABLE_TYPE")

            if parts:
                js_dat["Argument"]["PARTS"] = parts

            if "LOAD_CASE_NAMES" in js_dat["Argument"]:
                js_dat["Argument"]["TH_CASE_NAME"] = js_dat["Argument"].pop("LOAD_CASE_NAMES")

            js_dat["Argument"]["STEP"] = {
                "FROM": step_from,
                "TO": step_to,
                "STEPS": step_interval
            }
                
            js_dat["Argument"].pop("TABLE_NAME", None)
            ResultJSON = _changeUNITandGetDataText(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable_TEXT(table_type,ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def TH_PlaneStrainForce(th_case: list,keys=[],parts: list=["PartI", "PartJ","PartK","PartL"],
                    step_from: float = 0, step_to: float = 1, step_interval: int = 1,
                    components=['all'],
                    options: TableOptions = None):
            '''
            Fetches Time History PlaneStrainForce result tables.
            Args:
                keys (list/str): List of Plane Strain Element IDs.
                th_case (list): Time history load case names, e.g. ["Elcent"].
                step_from (float): Start time.
                step_to (float): End time.
                step_interval (int): Time interval (STEPS).
                components (list): Table components to include. Defaults to ['all'].
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"TH_PlaneStrainForce {_case2name(th_case)}"
            table_type = "TH_PLANE_STRAIN_FORCE"
            js_dat = _generate(table_type, keys, th_case, components, [], options)

            js_dat["Argument"]["TEXT_TYPE"] = js_dat["Argument"].pop("TABLE_TYPE")

            if parts:
                js_dat["Argument"]["PARTS"] = parts

            if "LOAD_CASE_NAMES" in js_dat["Argument"]:
                js_dat["Argument"]["TH_CASE_NAME"] = js_dat["Argument"].pop("LOAD_CASE_NAMES")

            js_dat["Argument"]["STEP"] = {
                "FROM": step_from,
                "TO": step_to,
                "STEPS": step_interval
            }
                
            js_dat["Argument"].pop("TABLE_NAME", None)
            ResultJSON = _changeUNITandGetDataText(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable_TEXT(table_type,ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def TH_SolidForce(th_case: list,keys=[],parts: list=["PartI", "PartJ","PartK","PartL","PartM","PartN","PartO","PartP"],
                    step_from: float = 0, step_to: float = 1, step_interval: int = 1,
                    components=['all'],
                    options: TableOptions = None):
            '''
            Fetches Time History SolidForce result tables.
            Args:
                keys (list/str): List of Solid Element IDs.
                th_case (list): Time history load case names, e.g. ["Elcent"].
                step_from (float): Start time.
                step_to (float): End time.
                step_interval (int): Time interval (STEPS).
                components (list): Table components to include. Defaults to ['all'].
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"TH_SolidForce {_case2name(th_case)}"
            table_type = "TH_SOLIDFORCE"
            js_dat = _generate(table_type, keys, th_case, components, [], options)

            js_dat["Argument"]["TEXT_TYPE"] = js_dat["Argument"].pop("TABLE_TYPE")

            if parts:
                js_dat["Argument"]["PARTS"] = parts

            if "LOAD_CASE_NAMES" in js_dat["Argument"]:
                js_dat["Argument"]["TH_CASE_NAME"] = js_dat["Argument"].pop("LOAD_CASE_NAMES")

            js_dat["Argument"]["STEP"] = {
                "FROM": step_from,
                "TO": step_to,
                "STEPS": step_interval
            }
                
            js_dat["Argument"].pop("TABLE_NAME", None)
            ResultJSON = _changeUNITandGetDataText(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable_TEXT(table_type,ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def TH_PlateForce(th_case: list,keys=[],parts: list=["PartI", "PartJ","PartK","PartL"],
                    step_from: float = 0, step_to: float = 1, step_interval: int = 1,
                    components=['all'],
                    options: TableOptions = None):
            '''
            Fetches Time History PlateForce result tables.
            Args:
                keys (list/str): List of Plate Element IDs.
                th_case (list): Time history load case names, e.g. ["Elcent"].
                step_from (float): Start time.
                step_to (float): End time.
                step_interval (int): Time interval (STEPS).
                components (list): Table components to include. Defaults to ['all'].
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"TH_PlateForce {_case2name(th_case)}"
            table_type = "TH_PLATEFORCE"
            js_dat = _generate(table_type, keys, th_case, components, [], options)

            js_dat["Argument"]["TEXT_TYPE"] = js_dat["Argument"].pop("TABLE_TYPE")

            if parts:
                js_dat["Argument"]["PARTS"] = parts

            if "LOAD_CASE_NAMES" in js_dat["Argument"]:
                js_dat["Argument"]["TH_CASE_NAME"] = js_dat["Argument"].pop("LOAD_CASE_NAMES")

            js_dat["Argument"]["STEP"] = {
                "FROM": step_from,
                "TO": step_to,
                "STEPS": step_interval
            }
                
            js_dat["Argument"].pop("TABLE_NAME", None)
            ResultJSON = _changeUNITandGetDataText(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable_TEXT(table_type,ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF
        
        @staticmethod
        def TH_WallForce(th_case: list,keys=[],parts: list=["PartI", "PartJ"],
                    step_from: float = 0, step_to: float = 1, step_interval: int = 1,
                    components=['all'],
                    options: TableOptions = None):
            '''
            Fetches Time History WallForce result tables.
            Args:
                keys (list/str): List of Wall Element IDs.
                th_case (list): Time history load case names, e.g. ["Elcent"].
                step_from (float): Start time.
                step_to (float): End time.
                step_interval (int): Time interval (STEPS).
                components (list): Table components to include. Defaults to ['all'].
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"TH_WallForce {_case2name(th_case)}"
            table_type = "TH_WALLFORCE"
            js_dat = _generate(table_type, keys, th_case, components, [], options)

            js_dat["Argument"]["TEXT_TYPE"] = js_dat["Argument"].pop("TABLE_TYPE")

            if parts:
                js_dat["Argument"]["PARTS"] = parts

            if "LOAD_CASE_NAMES" in js_dat["Argument"]:
                js_dat["Argument"]["TH_CASE_NAME"] = js_dat["Argument"].pop("LOAD_CASE_NAMES")

            js_dat["Argument"]["STEP"] = {
                "FROM": step_from,
                "TO": step_to,
                "STEPS": step_interval
            }
                
            js_dat["Argument"].pop("TABLE_NAME", None)
            ResultJSON = _changeUNITandGetDataText(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable_TEXT(table_type,ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF
        
        @staticmethod
        def TH_PlateUnitForce(th_case: list,keys=[],parts: list=["PartC","PartI", "PartJ","PartK","PartL"],
                    step_from: float = 0, step_to: float = 1, step_interval: int = 1,
                    components=['all'],
                    options: TableOptions = None):
            '''
            Fetches Time History PlateForce (Unit Length) result tables.
            Args:
                keys (list/str): List of Plate Element IDs.
                th_case (list): Time history load case names, e.g. ["Elcent"].
                step_from (float): Start time.
                step_to (float): End time.
                step_interval (int): Time interval (STEPS).
                components (list): Table components to include. Defaults to ['all'].
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"TH_PlateUnitForce {_case2name(th_case)}"
            table_type = "TH_PLATE_UNIT_FORCE"
            js_dat = _generate(table_type, keys, th_case, components, [], options)

            js_dat["Argument"]["TEXT_TYPE"] = js_dat["Argument"].pop("TABLE_TYPE")

            if parts:
                js_dat["Argument"]["PARTS"] = parts

            if "LOAD_CASE_NAMES" in js_dat["Argument"]:
                js_dat["Argument"]["TH_CASE_NAME"] = js_dat["Argument"].pop("LOAD_CASE_NAMES")

            js_dat["Argument"]["STEP"] = {
                "FROM": step_from,
                "TO": step_to,
                "STEPS": step_interval
            }
                
            js_dat["Argument"].pop("TABLE_NAME", None)
            ResultJSON = _changeUNITandGetDataText(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable_TEXT(table_type,ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def TH_BeamStress(th_case: list,keys=[],parts=["PartI", "PartJ"],
                    step_from: float = 0, step_to: float = 1, step_interval: int = 1,
                    components=['all'],
                    options: TableOptions = None):
            '''
            Fetches Time History BeamStress result tables.
            Args:
                keys (list/str): List of Beam Element IDs.
                th_case (list): Time history load case names, e.g. ["Elcent"].
                step_from (float): Start time.
                step_to (float): End time.
                step_interval (int): Time interval (STEPS).
                components (list): Table components to include. Defaults to ['all'].
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"TH_BeamStress {_case2name(th_case)}"
            table_type = "TH_BEAMSTRESS"
            js_dat = _generate(table_type, keys, th_case, components, [], options)

            js_dat["Argument"]["TEXT_TYPE"] = js_dat["Argument"].pop("TABLE_TYPE")
            
            if parts:
                js_dat["Argument"]["PARTS"] = parts
            
            if "LOAD_CASE_NAMES" in js_dat["Argument"]:
                js_dat["Argument"]["TH_CASE_NAME"] = js_dat["Argument"].pop("LOAD_CASE_NAMES")

            js_dat["Argument"]["STEP"] = {
                "FROM": step_from,
                "TO": step_to,
                "STEPS": step_interval
            }
                
            js_dat["Argument"].pop("TABLE_NAME", None)
            ResultJSON = _changeUNITandGetDataText(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable_TEXT(table_type,ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def TH_TrussStress(th_case: list,keys=[],
                    step_from: float = 0, step_to: float = 1, step_interval: int = 1,
                    components=['all'],
                    options: TableOptions = None):
            '''
            Fetches Time History TrussStress result tables.
            Args:
                keys (list/str): List of Truss Element IDs.
                th_case (list): Time history load case names, e.g. ["Elcent"].
                step_from (float): Start time.
                step_to (float): End time.
                step_interval (int): Time interval (STEPS).
                components (list): Table components to include. Defaults to ['all'].
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"TH_TrussStress {_case2name(th_case)}"
            table_type = "TH_TRUSSSTRESS"
            js_dat = _generate(table_type, keys, th_case, components, [], options)

            js_dat["Argument"]["TEXT_TYPE"] = js_dat["Argument"].pop("TABLE_TYPE")

            if "LOAD_CASE_NAMES" in js_dat["Argument"]:
                js_dat["Argument"]["TH_CASE_NAME"] = js_dat["Argument"].pop("LOAD_CASE_NAMES")

            js_dat["Argument"]["STEP"] = {
                "FROM": step_from,
                "TO": step_to,
                "STEPS": step_interval
            }
                
            js_dat["Argument"].pop("TABLE_NAME", None)
            ResultJSON = _changeUNITandGetDataText(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable_TEXT(table_type,ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF
        @staticmethod
        def TH_PlateStress(th_case: list,keys=[],parts: list=["PartC","PartI", "PartJ","PartK","PartL"],
                    step_from: float = 0, step_to: float = 1, step_interval: int = 1,
                    components=['all'],
                    options: TableOptions = None):
            '''
            Fetches Time History PlateStress result tables.
            Args:
                keys (list/str): List of Plate Element IDs.
                th_case (list): Time history load case names, e.g. ["Elcent"].
                step_from (float): Start time.
                step_to (float): End time.
                step_interval (int): Time interval (STEPS).
                components (list): Table components to include. Defaults to ['all'].
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"TH_PlateStress {_case2name(th_case)}"
            table_type = "TH_PLATESTRESS"
            js_dat = _generate(table_type, keys, th_case, components, [], options)

            js_dat["Argument"]["TEXT_TYPE"] = js_dat["Argument"].pop("TABLE_TYPE")

            if parts:
                js_dat["Argument"]["PARTS"] = parts

            if "LOAD_CASE_NAMES" in js_dat["Argument"]:
                js_dat["Argument"]["TH_CASE_NAME"] = js_dat["Argument"].pop("LOAD_CASE_NAMES")

            js_dat["Argument"]["STEP"] = {
                "FROM": step_from,
                "TO": step_to,
                "STEPS": step_interval
            }
                
            js_dat["Argument"].pop("TABLE_NAME", None)
            ResultJSON = _changeUNITandGetDataText(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable_TEXT(table_type,ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF
        
        @staticmethod
        def TH_PlaneStress_Stress(th_case: list,keys=[],parts: list=["PartC","PartI", "PartJ","PartK","PartL"],
                    step_from: float = 0, step_to: float = 1, step_interval: int = 1,
                    components=['all'],
                    options: TableOptions = None):
            '''
            Fetches Time History PlaneStress Stress result tables.
            Args:
                keys (list/str): List of PlaneStress Element IDs.
                th_case (list): Time history load case names, e.g. ["Elcent"].
                step_from (float): Start time.
                step_to (float): End time.
                step_interval (int): Time interval (STEPS).
                components (list): Table components to include. Defaults to ['all'].
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"TH_PlaneStress_Stress {_case2name(th_case)}"
            table_type = "TH_PLANESTRESS"
            js_dat = _generate(table_type, keys, th_case, components, [], options)

            js_dat["Argument"]["TEXT_TYPE"] = js_dat["Argument"].pop("TABLE_TYPE")

            if parts:
                js_dat["Argument"]["PARTS"] = parts

            if "LOAD_CASE_NAMES" in js_dat["Argument"]:
                js_dat["Argument"]["TH_CASE_NAME"] = js_dat["Argument"].pop("LOAD_CASE_NAMES")

            js_dat["Argument"]["STEP"] = {
                "FROM": step_from,
                "TO": step_to,
                "STEPS": step_interval
            }
                
            js_dat["Argument"].pop("TABLE_NAME", None)
            ResultJSON = _changeUNITandGetDataText(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable_TEXT(table_type,ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def TH_PlaneStrain_Stress(th_case: list,keys=[],parts: list=["PartC","PartI", "PartJ","PartK","PartL"],
                    step_from: float = 0, step_to: float = 1, step_interval: int = 1,
                    components=['all'],
                    options: TableOptions = None):
            '''
            Fetches Time History PlaneStrain Stress result tables.
            Args:
                keys (list/str): List of PlaneStrain Element IDs.
                th_case (list): Time history load case names, e.g. ["Elcent"].
                step_from (float): Start time.
                step_to (float): End time.
                step_interval (int): Time interval (STEPS).
                components (list): Table components to include. Defaults to ['all'].
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"TH_PlaneStrain_Stress {_case2name(th_case)}"
            table_type = "TH_PLANE_STRAIN_STRESS"
            js_dat = _generate(table_type, keys, th_case, components, [], options)

            js_dat["Argument"]["TEXT_TYPE"] = js_dat["Argument"].pop("TABLE_TYPE")

            if parts:
                js_dat["Argument"]["PARTS"] = parts

            if "LOAD_CASE_NAMES" in js_dat["Argument"]:
                js_dat["Argument"]["TH_CASE_NAME"] = js_dat["Argument"].pop("LOAD_CASE_NAMES")

            js_dat["Argument"]["STEP"] = {
                "FROM": step_from,
                "TO": step_to,
                "STEPS": step_interval
            }
                
            js_dat["Argument"].pop("TABLE_NAME", None)
            ResultJSON = _changeUNITandGetDataText(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable_TEXT(table_type,ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def TH_SolidStress(th_case: list,keys=[],parts: list=["PartI", "PartJ","PartK","PartL","PartM","PartN","PartO","PartP"],
                    step_from: float = 0, step_to: float = 1, step_interval: int = 1,
                    components=['all'],
                    options: TableOptions = None):
            '''
            Fetches Time History Solid Stress result tables.
            Args:
                keys (list/str): List of Solid Element IDs.
                th_case (list): Time history load case names, e.g. ["Elcent"].
                step_from (float): Start time.
                step_to (float): End time.
                step_interval (int): Time interval (STEPS).
                components (list): Table components to include. Defaults to ['all'].
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"TH_SolidStress {_case2name(th_case)}"
            table_type = "TH_SOLIDSTRESS"
            js_dat = _generate(table_type, keys, th_case, components, [], options)

            js_dat["Argument"]["TEXT_TYPE"] = js_dat["Argument"].pop("TABLE_TYPE")

            if parts:
                js_dat["Argument"]["PARTS"] = parts

            if "LOAD_CASE_NAMES" in js_dat["Argument"]:
                js_dat["Argument"]["TH_CASE_NAME"] = js_dat["Argument"].pop("LOAD_CASE_NAMES")

            js_dat["Argument"]["STEP"] = {
                "FROM": step_from,
                "TO": step_to,
                "STEPS": step_interval
            }
                
            js_dat["Argument"].pop("TABLE_NAME", None)
            ResultJSON = _changeUNITandGetDataText(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable_TEXT(table_type,ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF
        
        @staticmethod
        def VibrationModeShapes(modes, nodeIDs=[], output:_eigenOutputType="EigenVector",
                                 options:TableOptions=None):
            '''
            Fetches Vibration Mode Shape (Eigenvalue Mode) result tables.

            Args:
                modes (str/int/list): Mode number(s), e.g. 1, 'Mode 1', or ['Mode1','Mode2'].
                nodeIDs (list/str): List of Node IDs or a Structure Group Name.
                output (str): "EigenVector" | "Eigenvalue_Analysis" |
                               "Modal_Participation_Percent" | "Modal_Participation_Mass" |
                               "Modal_Participation_Factor" | "Modal_Direction_Factor"
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"VibrationModeShape {output}"

            table_type = "EIGENVALUEMODE"
            components = ["Node", "Mode", "UX", "UY", "UZ", "RX", "RY", "RZ"] 

            js_dat = _generate(table_type, nodeIDs, None, components, [], options)
            js_dat["Argument"]["MODES"] = _format_modes(modes)
            
            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT,options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable_Eigen(ResultJSON, output,options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def BucklingModeShapes(modes, nodeIDs=[], output:_bucklingOutputType="BucklingVector",
                                 options:TableOptions=None):
            '''
            Fetches Buckling Mode Shape result tables.

            Args:
                modes (str/int/list): Mode number(s), e.g. 1, 'Mode 1', or ['Mode1','Mode2'].
                nodeIDs (list/str): List of Node IDs or a Structure Group Name.
                output (str): "BucklingVector" | "Buckling_Analysis" |
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"VibrationModeShape {output}"

            table_type = "BUCKLINGMODE"
            components = ["Node", "Mode", "UX", "UY", "UZ", "RX", "RY", "RZ"] 

            js_dat = _generate(table_type, nodeIDs, None, components, [], options)
            js_dat["Argument"]["MODES"] = _format_modes(modes)
            
            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT,options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable_Eigen(ResultJSON, output,options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

    @staticmethod
    def IMAGE(ResultGraphic:ResultGraphic,location:str='',image_size:tuple = None,CS_StageName:str='',CS_StepIndex=2, bOutputImage=True):
        ''' 
        Capture Result Graphic in CIVIL NX   
            Result Graphic - ResultGraphic JSON (ResultGraphic.BeamDiagram())
            Location - image location
            Image height and width
            Construction stage Name (default = "") if desired
        '''
        if image_size==None: image_size=View.Image_Size
        _ERROR_MSG = ''
        json_body = {
                "Argument":{
                    "SET_MODE":"post",
                    "SET_HIDDEN":View.Hidden,
                    "HEIGHT":image_size[1],
                    "WIDTH":image_size[0],
                    "RESULT_GRAPHIC": ResultGraphic
                }
                }
        if View.Angle.__newH__ == True or View.Angle.__newV__ == True:
            json_body['Argument']['ANGLE'] = View.Angle._json()

        if View.Active.__default__ ==False:
            json_body['Argument']['ACTIVE'] = View.Active._json()

        if CS_StageName != '':
            json_body['Argument']['STAGE_NAME'] = CS_StageName

        json_body['Argument']['RESULT_GRAPHIC']['LOAD_CASE_COMB']['STEP_INDEX'] = CS_StepIndex
        
        resp = MidasAPI('POST','/view/CAPTURE',json_body)

        if "base64String" in resp:
            bs64_img = b64decode(resp["base64String"])

            if location:
                __img_file = open(location, 'wb')  # Open image file to save.
                __img_file.write(bs64_img)  # Decode and write data.
                __img_file.close()
            
            if bOutputImage:
                from PIL import Image
                from io import BytesIO

                # image = Image.new("RGB", image_size, "white")
                # buffer = BytesIO()
                # image.save(buffer)
                # buffer.seek(0)

                return Image.open(BytesIO(bs64_img))


                
                # return bs64_img
                return Image(temp_img)
            
            
        else:
            # ERROR IMAGE --------------
            try:
                _ERROR_MSG = resp['error']['message']
            except:
                _ERROR_MSG = "CANNOT RETRIEVE IMAGE. ERROR UNKNOWN"

            
            from PIL import Image, ImageDraw, ImageFont
            image = Image.new("RGB", image_size, "white")
            draw = ImageDraw.Draw(image)

            font = ImageFont.load_default()


            # Get text bounding box for centering
            bbox = draw.textbbox((0, 0), _ERROR_MSG, font=font)
            text_width = bbox[2] - bbox[0]
            text_height = bbox[3] - bbox[1]

            # Calculate centered position
            x = (image_size[0] - text_width) // 2
            y = (image_size[1] - text_height) // 2

            # Draw the text in black
            draw.text((x, y-15), "ERROR", fill="red", font=font)
            draw.text((x, y), _ERROR_MSG, fill="black", font=font)




            _RES_TYPE = ResultGraphic["CURRENT_MODE"]
            _LC_NAME = ResultGraphic["LOAD_CASE_COMB"]["NAME"]
            _IMG_DEF = f"Result Image    |    {_RES_TYPE}  for  {_LC_NAME}  load case.    |    Size  {image_size[0]}x{image_size[1]} px"

            draw.text((image_size[0]//2, image_size[1]-30), _IMG_DEF, fill="black", font=font,anchor='ms')

            if location:
                # Save the image
                image.save(location)
            
            if bOutputImage:
                return image           

        return resp